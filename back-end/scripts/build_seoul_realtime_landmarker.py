import argparse
import html
import json
import os
import re
import time
from pathlib import Path

import requests
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]

DEFAULT_INPUT_PATH = (
    PROJECT_ROOT
    / "data"
    / "raw"
    / "seoul_realtime_areas.xlsx"
)

DEFAULT_OUTPUT_PATH = (
    PROJECT_ROOT
    / "data"
    / "seoul_realtime_landmarks.json"
)

DEFAULT_UNRESOLVED_PATH = (
    PROJECT_ROOT
    / "data"
    / "seoul_realtime_landmarks_unresolved.json"
)

NAVER_LOCAL_SEARCH_URL = (
    "https://openapi.naver.com/v1/search/local.json"
)


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "서울시 실시간 인구 제공 장소 목록 XLSX를 읽어 "
            "네이버 지역 검색 API 기반 랜드마크 JSON으로 변환합니다."
        )
    )

    parser.add_argument(
        "--input-xlsx",
        type=Path,
        default=DEFAULT_INPUT_PATH,
        help=f"서울시 주요 장소 목록 XLSX 경로 (기본값: {DEFAULT_INPUT_PATH})",
    )

    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT_PATH,
        help=f"생성할 랜드마크 JSON 경로 (기본값: {DEFAULT_OUTPUT_PATH})",
    )

    parser.add_argument(
        "--unresolved-output",
        type=Path,
        default=DEFAULT_UNRESOLVED_PATH,
        help=(
            "좌표를 찾지 못했거나 검토가 필요한 장소 목록 경로 "
            f"(기본값: {DEFAULT_UNRESOLVED_PATH})"
        ),
    )

    parser.add_argument(
        "--code-column",
        default="AREA_CD",
        help="장소 코드 컬럼명 (기본값: AREA_CD)",
    )

    parser.add_argument(
        "--name-column",
        default="AREA_NM",
        help="장소명 컬럼명 (기본값: AREA_NM)",
    )

    parser.add_argument(
        "--request-delay",
        type=float,
        default=0.1,
        help="네이버 검색 요청 사이 대기 시간(초, 기본값: 0.1)",
    )

    parser.add_argument(
        "--allow-unresolved",
        action="store_true",
        help="좌표를 찾지 못한 장소가 있어도 성공 코드로 종료합니다.",
    )

    return parser.parse_args()


def get_required_environment(name: str) -> str:
    value = os.getenv(name)

    if not value:
        raise RuntimeError(
            f"환경변수 {name}이 설정되지 않았습니다."
        )

    return value


def normalize_header(value: object) -> str:
    return str(value or "").strip().upper()


def load_area_rows(
    input_path: Path,
    code_column: str,
    name_column: str,
) -> list[dict[str, str]]:
    if not input_path.exists():
        raise FileNotFoundError(
            f"입력 XLSX 파일을 찾을 수 없습니다: {input_path}"
        )

    workbook = load_workbook(
        input_path,
        read_only=True,
        data_only=True,
    )

    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)

    try:
        header_row = next(rows)
    except StopIteration as exc:
        raise RuntimeError(
            "XLSX 파일이 비어 있습니다."
        ) from exc

    header_index = {
        normalize_header(value): index
        for index, value in enumerate(header_row)
    }

    code_index = header_index.get(
        normalize_header(code_column)
    )
    name_index = header_index.get(
        normalize_header(name_column)
    )

    if code_index is None or name_index is None:
        available_columns = [
            str(value)
            for value in header_row
            if value not in (None, "")
        ]

        raise RuntimeError(
            "필수 컬럼을 찾지 못했습니다. "
            f"required=({code_column}, {name_column}), "
            f"available={available_columns}"
        )

    areas: list[dict[str, str]] = []

    for row in rows:
        code = row[code_index]
        name = row[name_index]

        if code in (None, "") or name in (None, ""):
            continue

        areas.append(
            {
                "area_code": str(code).strip(),
                "area_name": str(name).strip(),
            }
        )

    if not areas:
        raise RuntimeError(
            "XLSX에서 유효한 장소 코드·장소명을 찾지 못했습니다."
        )

    area_codes = [
        area["area_code"]
        for area in areas
    ]

    if len(area_codes) != len(set(area_codes)):
        raise RuntimeError(
            "XLSX에 중복된 장소 코드가 있습니다."
        )

    return areas


def strip_html(value: str) -> str:
    plain_text = re.sub(r"<[^>]+>", "", value)

    return html.unescape(plain_text).strip()

def search_coordinate(
    area_name: str,
    client_id: str,
    client_secret: str,
) -> dict | None:
    response = requests.get(
        NAVER_LOCAL_SEARCH_URL,
        params={
            "query": f"서울 {area_name}",
            "display": 5,
            "start": 1,
            "sort": "random",
        },
        headers={
            "X-Naver-Client-Id": client_id,
            "X-Naver-Client-Secret": client_secret,
        },
        timeout=15,
    )

    response.raise_for_status()

    payload = response.json()
    items = payload.get("items", [])

    if not items:
        return None

    first_item = items[0]

    try:
        longitude = float(first_item["mapx"])
        latitude = float(first_item["mapy"])
    except (KeyError, TypeError, ValueError):
        return None

    if abs(longitude) > 180:
        longitude /= 10_000_000

    if abs(latitude) > 90:
        latitude /= 10_000_000

    if not -90 <= latitude <= 90:
        return None

    if not -180 <= longitude <= 180:
        return None

    return {
        "latitude": latitude,
        "longitude": longitude,
        "matched_name": strip_html(
            str(first_item.get("title", ""))
        ),
        "matched_address": str(
            first_item.get("roadAddress")
            or first_item.get("address")
            or ""
        ).strip(),
    }


def write_json(
    output_path: Path,
    payload: object,
) -> None:
    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    output_path.write_text(
        json.dumps(
            payload,
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )


def main() -> None:
    arguments = parse_arguments()

    client_id = get_required_environment(
        "NAVER_SEARCH_CLIENT_ID"
    )
    client_secret = get_required_environment(
        "NAVER_SEARCH_CLIENT_SECRET"
    )

    areas = load_area_rows(
        input_path=arguments.input_xlsx,
        code_column=arguments.code_column,
        name_column=arguments.name_column,
    )

    landmarks: list[dict] = []
    unresolved: list[dict] = []

    for index, area in enumerate(areas, start=1):
        area_code = area["area_code"]
        area_name = area["area_name"]

        print(
            f"[{index}/{len(areas)}] "
            f"좌표 검색: {area_code} / {area_name}"
        )

        try:
            coordinate = search_coordinate(
                area_name=area_name,
                client_id=client_id,
                client_secret=client_secret,
            )
        except requests.RequestException as exc:
            unresolved.append(
                {
                    **area,
                    "reason": "naver_search_request_failed",
                    "detail": str(exc),
                }
            )
            time.sleep(arguments.request_delay)
            continue

        if coordinate is None:
            unresolved.append(
                {
                    **area,
                    "reason": "coordinate_not_found",
                }
            )
            time.sleep(arguments.request_delay)
            continue

        landmarks.append(
            {
                "area_code": area_code,
                "area_name": area_name,
                "latitude": coordinate["latitude"],
                "longitude": coordinate["longitude"],
            }
        )

        time.sleep(arguments.request_delay)

    landmarks.sort(
        key=lambda landmark: landmark["area_code"]
    )

    write_json(
        output_path=arguments.output,
        payload=landmarks,
    )

    write_json(
        output_path=arguments.unresolved_output,
        payload=unresolved,
    )

    print()
    print(f"완료된 랜드마크: {len(landmarks)}")
    print(f"검토 필요 랜드마크: {len(unresolved)}")
    print(f"출력 파일: {arguments.output}")
    print(f"검토 파일: {arguments.unresolved_output}")

    if unresolved and not arguments.allow_unresolved:
        raise SystemExit(
            "일부 장소의 좌표를 찾지 못했습니다. "
            "unresolved JSON을 검토한 뒤 다시 실행하세요."
        )


if __name__ == "__main__":
    main()